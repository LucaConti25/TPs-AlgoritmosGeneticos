import random
import sys
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Generacion aleatoria de un cromosoma
def generar_cromosoma(bits): 
    cromosoma =[random.randint(0,1) for _ in range(bits)]
    cromosoma = ''.join(map(str, cromosoma))
    return cromosoma

# Generacion de una poblacion inicial de cant_cromo individuos
def genera_poblacion_inicial(cant_cromo, bits): 
    return [generar_cromosoma(bits) for i in range(cant_cromo)]

# Calcula el valor de la funcion objetivo para toda la poblacion
def calcula_obj(poblacion,coef): 
    return[((int(cromosoma,2)/coef)**2) for cromosoma in poblacion]

# Calcula el fitness para toda la poblacion
def calcula_fitness(valores_objs, sumatoria): 
   return [i/sumatoria for i in valores_objs]

# Busca el valor maximo de la poblacion y el cromosoma correspondiente a dicho valor
def busca_maximo(poblacion, coef): 
    max = 0
    ganador = 0
    for cromo in poblacion:
        val = int(cromo,2)
        if (((val/coef)**2) > max):
            ganador = cromo
            max = ((val/coef)**2)
    return ganador

# Realiza el metodo de seleccion por Torneo   
def torneo(cromosomas, fit): 
    seleccionados = []
    tam = len(fit)
    for _ in range(tam):
        idA = random.randint(0,tam-1)
        idB = random.randint(0,tam-1)
        # Si el Fitness de A es mayor se elige A
        if (fit[idA] > fit[idB]): seleccionados.append(cromosomas[idA]) 
        # Si el fitness de B es mayor o igual se elige B
        else: seleccionados.append(cromosomas[idB]) 
    
    # Se devuelven los cromosomas seleccionados para la cruza
    return seleccionados 

# Realiza el metodo de seleccion por Ruleta
def ruleta(cromosomas, fit,cant): 
    seleccionados = []
    ruleta = [] 
    # Se genera una ruleta donde la cantidad de espacios asignada a cada cromosoma es proporcional a su fitness
    for i in range(len(fit)): 
        # A cada cromosoma le corresponde minimo 1 espacio
        val = max(int(fit[i]*100),1) 
        for _ in range(val): ruleta.append(i)
    
    # Si quedaron espacios sin asignar se asignan a mano
    while(len(ruleta) < 100): ruleta.append(len(fit)-1) 
    
    for _ in range(cant):
        salida = random.randint(0,99)
        # Se eligen los cromosomas
        seleccionados.append(cromosomas[ruleta[salida]]) 
    
    # Se devuelven los cromosomas seleccionados para la cruza
    return seleccionados 

# Realiza el metodo de seleccion por Elitismo y luego se llama al metodo de seleccion por Ruleta
def ruleta_elite(cromosomas,fit): 
    seleccionados = []
    tam = len(fit)
    elite = []
    for i in range(tam): elite.append([fit[i],i])
    # Se ordenan los cromosomas de acuerdo a su fitness
    elite.sort() 
    # Se elige el cromosoma con el mayor valor de fitness
    idA = elite[tam-1][1] 
    # Se elige el cromosoma con el segundo mayor valor de fitness
    idB = elite[tam-2][1] 
    seleccionados.append(cromosomas[idA])
    seleccionados.append(cromosomas[idB])
    # Se realiza el metodo de Ruleta descontando los 2 individuos elegidos por elitismo 
    seleccion_ruleta = ruleta(cromosomas,fit,tam-2) 
    # Se devuelven los cromosomas seleccionados para la cruza
    return seleccionados + seleccion_ruleta 

# Se realiza la cruza de los cromosomas seleccionados
def cruza(seleccionados,prob_cross,prob_mut,elitismo): 
    descendencia = []
    for i in range(0, len(seleccionados) - 1,2):
        if elitismo == 1 and i==0:
            # print("ENTRO")
            descendencia.append(seleccionados[i])
            descendencia.append(seleccionados[i+1])
            continue
        prob = random.random()
        # Si el numero generado es menor o igual a la probabilidad de cruza la misma se realiza
        if (prob <= prob_cross): 
            # Se genera de forma aleatoria el punto de corte
            corte = random.randint(0,29) 
            padre1 = seleccionados[i]
            padre2 = seleccionados[i+1]
            
            # Se genera el primer hijo al concatenar el primer padre hasta el corte con el segundo padre despues del corte
            hijo1 = padre1[:corte] + padre2[corte:] 
            # Se genera el primer hijo al concatenar el segundo padre hasta el corte con el primer padre despues del corte
            hijo2 = padre2[:corte] + padre1[corte:] 
            
            # Se guarda el primer hijo generado
            descendencia.append(hijo1) 
            # Se guarda el segundo hijo generado
            descendencia.append(hijo2) 
        else:
            # En caso de no realizarse la cruza se guarda el primer padre sin realizar ningun cambio
            descendencia.append(seleccionados[i]) 
            # En caso de no realizarse la cruza se guarda el segundo padre sin realizar ningun cambio
            descendencia.append(seleccionados[i+1]) 
    # Se realiza la mutacion de la poblacion
    if elitismo==1:
        descendencia = descendencia[0:2] + mutacion(descendencia[2:len(descendencia)],prob_mut)
    else:
        descendencia = mutacion(descendencia,prob_mut) 
    # Se devuelve la descendencia generada
    return descendencia 

# Se realiza la mutacion de los cromosomas
def mutacion(hijos,prob_mut): 
    hijos_nuevos = []
    for hijo in hijos:
        hijo_nuevo = ""
        prob = random.random()
        # Si el numero generado es menor o igual a la probabilidad de mutacion la misma se realiza
        if (prob <= prob_mut): 
            # Se genera el limite inferior
            lim_inf = random.randint(0,29) 
            # Se genera el limite superior
            lim_sup = random.randint(0,29) 

            # Se intercambian los valores si el limite inferior es mayor al superior          
            if (lim_inf > lim_sup): lim_inf, lim_sup = lim_sup, lim_inf  
            
            # Se invierte el segmento generado
            segmento_inv = hijo[lim_inf:lim_sup+1][::-1] 
            # Se genera el individuo al concatenar el nuevo segmento invertido con los segmentos viejos
            hijo_nuevo = hijo[:lim_inf] + segmento_inv + hijo[lim_sup+1:] 
            
            hijos_nuevos.append(hijo_nuevo) 
        # En caso de no realizarse la mutacion se guarda el individuo sin realizarse ningun cambio          
        else: hijos_nuevos.append(hijo) 
    # Se devulve la poblacion mutada
    return hijos_nuevos 

# Se generan las graficas de valor maximo, valor minimo y valor promedio
def hacer_grafica(maximos,minimos,promedios,cant_individuos,cant_corridas): 
    numero_corrida = [i+1 for i in range(cant_corridas)] 
    for i in range(cant_corridas):
        # Se calculan los promedios
        promedios[i] = promedios[i]/cant_individuos 
    # Se determina el tamaño de la grafica
    fig = plt.figure(figsize=(15, 10))
    # Se determina la grid para las graficas 
    gs = gridspec.GridSpec(2,2)
    ax1 = fig.add_subplot(gs[0,0])
    # Se pasan los datos generados y se selecciona varios parametros
    ax1.plot(numero_corrida,maximos,linestyle='-',label='Valor Maximo') 
    # Se indica el titulo de la grafica
    ax1.set_title('Maximo por corrida') 
    # Se indica el nombre del eje X
    ax1.set_xlabel('Numero Corrida') 
    # Se indica el nombre del eje Y
    ax1.set_ylabel('Valor Maximo') 
    # Se indica que la leyenda con el nombre de cada grafica debe aparecer
    ax1.legend() 

    # Se realiza el mismo procedimiento para el resto de graficas
    ax2 = fig.add_subplot(gs[0,1])
    ax2.plot(numero_corrida,minimos,linestyle='-',label='Valor Minimo')
    ax2.set_title('Minimo por corrida')
    ax2.set_xlabel('Numero Corrida')
    ax2.set_ylabel('Valor Minimo')
    ax2.legend()

    ax3 = fig.add_subplot(gs[1,:])
    ax3.plot(numero_corrida,promedios,linestyle='-',label='Valor Promedio')
    ax3.set_title('Promedio por corrida')
    ax3.set_xlabel('Numero Corrida')
    ax3.set_ylabel('Valor Promedio')
    ax3.legend()
    plt.tight_layout()
    plt.show()

# Se genera un archivo excel con los datos generados
def hacer_tabla(cromMax,maximos,minimos,promedios,cant_individuos,cant_corridas): 
    for i in range(cant_corridas):
        # Se calculan los promedios
        promedios[i] = promedios[i]/cant_individuos 
    # Se genera un DataFrame con los datos generados y se indica el nombre de las columnas
    df = pd.DataFrame(list(zip(cromMax,maximos,minimos,promedios)),columns=['Cromosoma Maximo','Valor Maximo','Valor Minimo','Promedio']) 
    wb = Workbook()
    ws = wb.active
    cont = 0
    # Se recorren las filas de la tabla
    for r in dataframe_to_rows(df, index=True, header=True): 
        if (cont == 0):
            # Se indica el nombre de la primer columna
            r[0] = "Generacion" 
        if (cont >= 2):
            # Se actualiza el indice para que el mismo sea 1-index
            r[0] = r[0]+1 
        # Se inserta la fila en el archivo excel
        ws.append(r) 
        cont += 1
    for cell in ws['A'] + ws[1]:
        # Se aplico estilo a las celdas
        cell.style = 'Pandas' 
    # Se guarda el archivo excel
    wb.save("pandas_openpyxl.xlsx") 
    
def main():
    # Se realiza el ingreso de los parametros
    if (len(sys.argv) != 11 or sys.argv[1] != "-c" or sys.argv[3] != "-m" or sys.argv[5] != "-n" or sys.argv[7] != "-o" or sys.argv[9] != "-g"):
        print("Este programa se utiliza con python tp1.py -c <prob_cross> -m <prob_mut> -n <cant_individuos> -o <metodo_seleccion> -g <cant_generaciones>")
        sys.exit()
    # Se inicializan las variables
    prob_cross = float(sys.argv[2])
    prob_mut = float(sys.argv[4])
    cant_individuos = int(sys.argv[6])
    op = int(sys.argv[8])
    generaciones = int(sys.argv[10])
    coef = (2**30) - 1
    cant_bits = 30
    max_por_ciclo = []
    valormin_por_ciclo = []
    valormax_por_ciclo = []
    sumas_obj_por_ciclo = []

    # Se genera la poblacion inicial de forma aleatoria
    poblacion = genera_poblacion_inicial(cant_individuos, cant_bits) 
    
    for i in range(generaciones):
        # Se calcula los valores de la funcion objetivo para la poblacion
        valores_func_obj = calcula_obj(poblacion, coef) 
        # Se guarda el valor minimo de la poblacion
        valormin_por_ciclo.append(min(valores_func_obj)) 
        # Se guarda el valor maximo de la poblacion
        valormax_por_ciclo.append(max(valores_func_obj)) 
        # Se guarda la suma de los valores de la poblacion
        sumas_obj_por_ciclo.append(sum(valores_func_obj)) 
        # Se guarda el cromosoma correspondiente al valor maximo
        max_por_ciclo.append(busca_maximo(poblacion, coef)) 
        sumatoria = sum(valores_func_obj)
        # Se calcula el fitness para toda la poblacion
        fitness = calcula_fitness(valores_func_obj, sumatoria) 
        # Se realiza el metodo de seleccion por Ruleta si el mismo fue seleccionado
        if (op == 0): seleccionados = ruleta(poblacion, fitness,10) 
        # Se realiza el metodo de seleccion por Torneo si el mismo fue seleccionado
        elif (op == 1): seleccionados = torneo(poblacion,fitness) 
        # Se realiza el metodo de seleccion por Elitismo+Ruleta si el mismo fue seleccionado
        elif (op == 2): seleccionados = ruleta_elite(poblacion,fitness) 
        else:
            print("El metodo de seleccion elegido no es valido")
            print("Opciones:")
            print("-0: Ruleta")
            print("-1: Torneo")
            print("-2: Ruleta + Elitismo")
            sys.exit()
        # Se realiza la cruza para la siguiente poblacion
        poblacion = cruza(seleccionados,prob_cross,prob_mut,op==2) 
    # Se imprimen en pantalla los datos de cada generacion
    for i in range(generaciones): 
        print("========================================================================")
        print(f"Valor maximo en poblacion {i}: {valormax_por_ciclo[i]} con el cromosoma {max_por_ciclo[i]}") 
        print(f"Valor minimo en poblacion {i}: {valormin_por_ciclo[i]}")
        print(f"Promedio de poblacion {i}: {sumas_obj_por_ciclo[i]/cant_individuos}")
        print("========================================================================")
    hacer_tabla(max_por_ciclo,valormax_por_ciclo,valormin_por_ciclo,sumas_obj_por_ciclo,cant_individuos,generaciones) # Se genera una archivo excel con los datos
    hacer_grafica(valormax_por_ciclo,valormin_por_ciclo,sumas_obj_por_ciclo,cant_individuos,generaciones) # Se generan graficas con los datos
   
main()